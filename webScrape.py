import pandas as pd
import urllib.request
import bs4 as bs
from openpyxl import load_workbook
from pandas import datetime

news = "https://www.breitbart.com/"
newsFile = urllib.request.urlopen(news).read()
soup = bs.BeautifulSoup(newsFile, 'html.parser')


sauce = soup.find(class_="top_article_main")
headline = sauce.h2.a.text
desc = sauce.div.p.text
mainheadline = (headline + " | " + desc)
main = pd.Series(mainheadline, name='MainHeadines')


sauce2 = soup.find("section", {"class": "featured_side"})
post = sauce2.find_all(class_="post")
features = [posts.find("h2").text for posts in post]
feature = pd.Series(features, name='FeaturedHeadlines')


sauce3 = soup.find("section", {"id": "BBTrendNow"})
trend = sauce3.find_all('li')
trending = [trends.find("a").text for trends in trend]
trends = pd.Series(trending, name='TrendingHeadlines')


df = pd.concat([main,feature,trends], axis=1,)
df.insert(0, 'TimeStamp', pd.datetime.now().replace(microsecond=0))
#Use this command and comment the one below if you have not initialized the data yet.
# df.to_excel('Breitbart.xlsx',index=False, header=True)

book = load_workbook('Breitbart.xlsx')
writer = pd.ExcelWriter('Breitbart.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = {ws.title: ws for ws in book.worksheets}

for sheetname in writer.sheets:
    df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row +2, index = False,header= False)

writer.save()
df
